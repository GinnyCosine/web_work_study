from selenium import webdriver
from selenium.webdriver.support.ui import Select
import pandas as pd
import numpy as np
import requests
import openpyxl
import csv
import time
import shutil
import os

# download path
f = open("path.txt", "r")
download = f.read()

# clean
for fname in range(0,7):
    if os.path.exists(str(fname)+".csv"):
        os.remove(str(fname)+".csv")

if os.path.exists(download + "StockList.csv"):
    os.remove(download + "StockList.csv")        
for sel in range(1,7):
    f_download = download + "StockList (" + str(sel) + ").csv"
    if os.path.exists(f_download):
        os.remove(f_download)
        
if os.path.exists('./result.xlsx'):
    
    driver = webdriver.Chrome()
    url = 'https://goodinfo.tw/StockInfo/StockList.asp?SHEET=%E8%82%A1%E5%88%A9%E6%94%BF%E7%AD%96&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E5%90%88%E8%A8%88%E8%82%A1%E5%88%A9+%28%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6%29%40%40%E5%90%88%E8%A8%88%E8%82%A1%E5%88%A9%40%40%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6'
    driver.get(url)
    time.sleep(1)
    for sel in range(0,7):
        select2 = Select(driver.find_element_by_id("selRANK"))
        select2.select_by_value(str(sel))
        time.sleep(5)
        btn = driver.find_element_by_css_selector("input[value='匯出CSV']")
        btn.click()
    while os.path.exists(download + 'StockList (6).csv') == False:
        time.sleep(0.1)
    shutil.move(download + "StockList.csv", './0.csv')
    for sel in range(1,7):
        shutil.move(download + "StockList (" + str(sel) + ").csv", str(sel) + ".csv")

    driver.close()
    ids = []
    prices = []
    changes = []
    
    for fname in range(0,7):
        df = pd.read_csv('./'+ str(fname) +'.csv')
        for nb in range(0, df.shape[0]):
            row = df.iloc[nb]
            ids.append(row['代號'].strip('="'))
            prices.append(row['成交'])
            changes.append(row['漲跌價'])
        
    fn = 'result.xlsx'
    wb = openpyxl.load_workbook(fn)
    wb.active = 0
    ws = wb.active
    row_cnt = ws.max_row
    for row in range(2, row_cnt + 1):
        id_ = ws['A'+str(row)].value
        if id_ in ids:
            idx = ids.index(id_)
            if ws['L'+str(row)].value != '-':
                avg = float(ws['L'+str(row)].value)
                ws['E'+str(row)].value = round(avg/prices[idx]*100, 6)
                ws['C'+str(row)].value = prices[idx]
                ws['D'+str(row)].value = changes[idx]
            else:
                ws['C'+str(row)].value = prices[idx]
                ws['D'+str(row)].value = changes[idx]
    wb.save(fn)
else:
    driver = webdriver.Chrome()
    url = 'https://goodinfo.tw/StockInfo/StockList.asp?SHEET=%E8%82%A1%E5%88%A9%E6%94%BF%E7%AD%96&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E5%90%88%E8%A8%88%E8%82%A1%E5%88%A9+%28%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6%29%40%40%E5%90%88%E8%A8%88%E8%82%A1%E5%88%A9%40%40%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6'
    driver.get(url)
    time.sleep(1)
    select1 = Select(driver.find_element_by_id("selSHEET"))
    select1.select_by_value('股利政策_近N年股利一覽')
    time.sleep(1)
    select1 = Select(driver.find_element_by_id("selSHEET2"))
    select1.select_by_value('現金股利')
    time.sleep(1)
    for sel in range(0,7):
        select2 = Select(driver.find_element_by_id("selRANK"))
        select2.select_by_value(str(sel))
        time.sleep(5)
        btn = driver.find_element_by_css_selector("input[value='匯出CSV']")
        btn.click()
    while os.path.exists(download + 'StockList (6).csv') == False:
        time.sleep(0.1)
    shutil.move(download + "StockList.csv", './0.csv')
    for sel in range(1,7):
        shutil.move(download + "StockList (" + str(sel) + ").csv", str(sel) + ".csv")

    driver.close()

    ids = []
    names = []
    prices = []
    changes = []
    yields = []
    dividends = {
        '2016': [],
        '2017': [],
        '2018': [],
        '2019': [],
        '2020': []
    }
    avg = []
    low = []
    mid = []
    high = []
    eps = []
    for fname in range(0,7):
        df = pd.read_csv('./'+ str(fname) +'.csv')
        for nb in range(0, df.shape[0]):
            row = df.iloc[nb]
            ids.append(row['代號'].strip('="'))
            names.append(row['名稱'])
            prices.append(row['成交'])
            changes.append(row['漲跌價'])
            if row['EPS平均'] != '-':
                eps.append(float(row['EPS平均']))
            else:
                eps.append('-')
            flag = 0
            total = 0
            div = 0
            for year in range(2020, 2015, -1):
                if flag == 1:
                    dividends[str(year)].append('-')
                elif pd.isnull(row[str(year)+'發放']) is True:
                    flag = 1
                    dividends[str(year)].append('-')
                else:
                    dividends[str(year)].append(row[str(year)+'發放'])
                    total += float(row[str(year)+'發放'])
                    div += 1
            if div > 0:
                a = total/div
                avg.append(round(a,2))
                yields.append(round(a/float(row['成交'])*100,6))
                low.append(round(a*16,3))
                mid.append(round(a*20,3))
                high.append(round(a*32,3))
            else:
                avg.append('-')
                yields.append('-')
                low.append('-')
                mid.append('-')
                high.append('-')
    data = {
        '證券編號': ids,
        '證券名稱': names,
        '成交價': prices,
        '漲跌價': changes,
        '殖利率(%)': yields,
        'EPS平均': eps,
        '2016股利': dividends['2016'],
        '2017股利': dividends['2017'],
        '2018股利': dividends['2018'],
        '2019股利': dividends['2019'],
        '2020股利': dividends['2020'],
        '平均股利': avg,
        '便宜價格': low,
        '合理價格': mid,
        '偏高價格': high
    }
    df = pd.DataFrame(data)
    df.to_excel('result.xlsx', index=False)

for fname in range(0,7):
    if os.path.exists(str(fname)+".csv"):
        os.remove(str(fname)+".csv")

print('Finish !')