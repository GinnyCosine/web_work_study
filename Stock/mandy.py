from selenium import webdriver
from selenium.webdriver.support.ui import Select
import pandas as pd
import numpy as np
import requests
import openpyxl
import csv
import time
import shutil
import os
import datetime

# time
today = datetime.date.today()
year_ = today.year
month_ = today.month

# download path
f = open("path.txt", "r")
download = f.read()
f.close()

# clean
for fname in range(0,7):
    if os.path.exists(str(fname)+".csv"):
        os.remove(str(fname)+".csv")

if os.path.exists(download + "StockList.csv"):
    os.remove(download + "StockList.csv")        
for sel in range(1,7):
    f_download = download + "StockList (" + str(sel) + ").csv"
    if os.path.exists(f_download):
        os.remove(f_download)        

driver = webdriver.Chrome()
if os.path.exists('./mandy.xlsx'):
    url = 'https://goodinfo.tw/StockInfo/StockList.asp?SHEET=%E8%82%A1%E5%88%A9%E6%94%BF%E7%AD%96&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E5%90%88%E8%A8%88%E8%82%A1%E5%88%A9+%28%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6%29%40%40%E5%90%88%E8%A8%88%E8%82%A1%E5%88%A9%40%40%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6'
    driver.get(url)
    for sel in range(0,7):
        select2 = Select(driver.find_element_by_id("selRANK"))
        select2.select_by_value(str(sel))
        time.sleep(6)
        btn = driver.find_element_by_css_selector("input[value='匯出CSV']")
        btn.click()
    while os.path.exists(download + 'StockList (6).csv') == False:
        time.sleep(0.1)
    shutil.move(download + "StockList.csv", './0.csv')
    for sel in range(1,7):
        shutil.move(download + "StockList (" + str(sel) + ").csv", str(sel) + ".csv")


    driver.close()
    ids = []
    prices = []
    changes = []
    
    for fname in range(0,7):
        df = pd.read_csv('./'+ str(fname) +'.csv')
        for nb in range(0, df.shape[0]):
            row = df.iloc[nb]
            ids.append(row['代號'].strip('="'))
            prices.append(row['成交'])
            changes.append(row['漲跌價'])
        
    fn = 'mandy.xlsx'
    wb = openpyxl.load_workbook(fn)
    wb.active = 0
    ws = wb.active
    row_cnt = ws.max_row
    for row in range(2, row_cnt + 1):
        id_ = ws['A'+str(row)].value
        idx = ids.index(id_)
        if ws['O'+str(row)].value != '-':
            avg = float(ws['L'+str(row)].value)
            ws['H'+str(row)].value = round(avg/prices[idx]*100, 6)
            ws['C'+str(row)].value = prices[idx]
            ws['D'+str(row)].value = changes[idx]
        else:
            ws['C'+str(row)].value = prices[idx]
            ws['D'+str(row)].value = changes[idx]
    wb.save(fn)
else:
    # 股利
    url = 'https://goodinfo.tw/StockInfo/StockList.asp?SHEET=%E8%82%A1%E5%88%A9%E6%94%BF%E7%AD%96&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E5%90%88%E8%A8%88%E8%82%A1%E5%88%A9+%28%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6%29%40%40%E5%90%88%E8%A8%88%E8%82%A1%E5%88%A9%40%40%E6%9C%80%E6%96%B0%E5%B9%B4%E5%BA%A6'
    driver.get(url)
    select1 = Select(driver.find_element_by_id("selSHEET"))
    select1.select_by_value('股利政策_近N年股利一覽')
    time.sleep(5)
    for sel in range(0,7):
        select2 = Select(driver.find_element_by_id("selRANK"))
        select2.select_by_value(str(sel))
        time.sleep(6)
        btn = driver.find_element_by_css_selector("input[value='匯出CSV']")
        btn.click()
    while os.path.exists(download + 'StockList (6).csv') == False:
        time.sleep(0.1)
    shutil.move(download + "StockList.csv", './0.csv')
    for sel in range(1,7):
        shutil.move(download + "StockList (" + str(sel) + ").csv", str(sel) + ".csv")

    ids = []
    names = []
    prices = []
    changes = []
    yields = []
    avg = []
    low = []
    mid = []
    high = []
    eps = []
    dividends = {}
    year_start = year_

    for fname in range(0,7):
        df = pd.read_csv('./'+ str(fname) +'.csv')
        if fname == 0:
            t = 0
            while (str(year_start)+'發放') not in df.columns:
                year_start -= 1
            for year in range(year_start, year_start-5, -1):
                dividends[str(year)] = []
        for nb in range(0, df.shape[0]):
            row = df.iloc[nb]
            ids.append(row['代號'].strip('="'))
            names.append(row['名稱'])
            prices.append(float(row['成交']))
            changes.append(float(row['漲跌價']))
            if row['EPS平均'] != '-':
                eps.append(float(row['EPS平均']))
            else:
                eps.append('-')
            total = 0
            div = 0
            for year in range(year_start, year_start-5, -1):
                if pd.isnull(row[str(year)+'發放']) is True:
                    dividends[str(year)].append('-')
                else:
                    dividends[str(year)].append(row[str(year)+'發放'])
                    total += float(row[str(year)+'發放'])
                    div += 1
            if div > 0:
                a = total/div
                avg.append(round(a,2))
                yields.append(round(a/float(row['成交'])*100,6))
                low.append(round(a*16,3))
                mid.append(round(a*20,3))
                high.append(round(a*32,3))
            else:
                avg.append('-')
                yields.append('-')
                low.append('-')
                mid.append('-')
                high.append('-')
    data = {
        '證券編號': ids,
        '證券名稱': names,
        '成交價': prices,
        '漲跌價': changes,
        '殖利率(%)': yields,
        'EPS平均': eps,
        '平均股利': avg,
        '便宜價格': low,
        '合理價格': mid,
        '偏高價格': high
    }
    for year in range(year_start, year_start-5, -1):
        data[str(year)+'股利'] = dividends[str(year)]
    df1 = pd.DataFrame(data)

    # 本益比
    url = 'https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E6%88%90%E4%BA%A4%E5%83%B9+%28%E9%AB%98%E2%86%92%E4%BD%8E%29%40%40%E6%88%90%E4%BA%A4%E5%83%B9%40%40%E7%94%B1%E9%AB%98%E2%86%92%E4%BD%8E'
    time.sleep(2)
    driver.get(url)
    for sel in range(0,7):
        select2 = Select(driver.find_element_by_id("selRANK"))
        select2.select_by_value(str(sel))
        time.sleep(6)
        btn = driver.find_element_by_css_selector("input[value='匯出CSV']")
        btn.click()
    while os.path.exists(download + 'StockList (6).csv') == False:
        time.sleep(0.1)
    shutil.move(download + "StockList.csv", './0.csv')
    for sel in range(1,7):
        shutil.move(download + "StockList (" + str(sel) + ").csv", str(sel) + ".csv")

    ids = []
    PERs = []

    for fname in range(0,7):
        df = pd.read_csv('./'+ str(fname) +'.csv')
        for nb in range(0, df.shape[0]):
            row = df.iloc[nb]
            ids.append(row['代號'].strip('="'))
            if pd.isnull(row['PER']):
                PERs.append('-')
            else:
                PERs.append(row['PER'])
            
    data = {
        '證券編號': ids,
        '本益比': PERs
    }
    df2 = pd.DataFrame(data)

    # 淨利率
    url = 'https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E5%96%AE%E5%AD%A3%E7%87%9F%E6%A5%AD%E5%88%A9%E7%9B%8A%E6%9C%80%E9%AB%98%40%40%E7%87%9F%E6%A5%AD%E5%88%A9%E7%9B%8A%40%40%E5%96%AE%E5%AD%A3%E7%87%9F%E6%A5%AD%E5%88%A9%E7%9B%8A%E6%9C%80%E9%AB%98'
    time.sleep(2)
    driver.get(url)
    select1 = Select(driver.find_element_by_id("selSHEET2"))
    select1.select_by_value('稅後淨利率(%)')
    time.sleep(5)
    for sel in range(0,6):
        select2 = Select(driver.find_element_by_id("selRANK"))
        select2.select_by_value(str(sel))
        time.sleep(6)
        btn = driver.find_element_by_css_selector("input[value='匯出CSV']")
        btn.click()
    while os.path.exists(download + 'StockList (5).csv') == False:
        time.sleep(0.1)
    shutil.move(download + "StockList.csv", './0.csv')
    for sel in range(1,6):
        shutil.move(download + "StockList (" + str(sel) + ").csv", str(sel) + ".csv")

    ids = []
    profits = []
    
    season = int((month_-1)/3)
    year = year_
    if season == 0:
        year -= 1
        season = 4

    for fname in range(0,6):
        df = pd.read_csv('./'+ str(fname) +'.csv')
        for nb in range(0, df.shape[0]):
            row = df.iloc[nb]
            ids.append(row['代號'].strip('="'))
            total = 0
            times = 0
            while times < 4:
                title = str(year)[2:] + 'Q' + str(season) + '淨利(%)'
                if title in df.columns:
                    total += float(row[title])
                    times += 1
                else:
                    season -= 1
                    if season == 0:
                        year -= 1
                        season = 4
            profits.append(round(total/4, 2))

    data = {
        '證券編號': ids,
        '平均淨利率(%)-近四季': profits
    }
    df3 = pd.DataFrame(data)

    # 毛利率
    url = 'https://goodinfo.tw/StockInfo/StockList.asp?RPT_TIME=&MARKET_CAT=%E7%86%B1%E9%96%80%E6%8E%92%E8%A1%8C&INDUSTRY_CAT=%E5%96%AE%E5%AD%A3%E6%AF%9B%E5%88%A9%E6%9C%80%E9%AB%98%40%40%E7%87%9F%E6%A5%AD%E6%AF%9B%E5%88%A9%40%40%E5%96%AE%E5%AD%A3%E6%AF%9B%E5%88%A9%E6%9C%80%E9%AB%98'
    time.sleep(2)
    driver.get(url)
    select1 = Select(driver.find_element_by_id("selSHEET"))
    select1.select_by_value('季獲利能力_近N季一覽')
    time.sleep(5)
    select1 = Select(driver.find_element_by_id("selSHEET2"))
    select1.select_by_value('營業毛利率(%)')
    time.sleep(5)
    for sel in range(0,6):
        select2 = Select(driver.find_element_by_id("selRANK"))
        select2.select_by_value(str(sel))
        time.sleep(6)
        btn = driver.find_element_by_css_selector("input[value='匯出CSV']")
        btn.click()
    time.sleep(2)
    while os.path.exists(download + 'StockList (5).csv') == False:
        time.sleep(0.1)
    shutil.move(download + "StockList.csv", './0.csv')
    for sel in range(1,6):
        shutil.move(download + "StockList (" + str(sel) + ").csv", str(sel) + ".csv")

    driver.close()
    ids = []
    profits = []
    
    season = int((month_-1)/3)
    year = year_
    if season == 0:
        year -= 1
        season = 4

    for fname in range(0,6):
        df = pd.read_csv('./'+ str(fname) +'.csv')
        for nb in range(0, df.shape[0]):
            row = df.iloc[nb]
            ids.append(row['代號'].strip('="'))
            total = 0
            times = 0
            while times < 4:
                title = str(year)[2:] + 'Q' + str(season) + '毛利(%)'
                if title in df.columns:
                    total += float(row[title])
                    times += 1
                else:
                    season -= 1
                    if season == 0:
                        year -= 1
                        season = 4
            profits.append(round(total/4, 2))

    data = {
        '證券編號': ids,
        '平均毛利率(%)-近四季': profits
    }
    df4 = pd.DataFrame(data)

    df = pd.merge(df3, df4, on='證券編號', how="outer")
    df = pd.merge(df2, df, on='證券編號', how="outer")
    df = pd.merge(df1, df, on='證券編號', how="outer")
    df = df[['證券編號','證券名稱','成交價','漲跌價','平均淨利率(%)-近四季','平均毛利率(%)-近四季','本益比','殖利率(%)','EPS平均','平均股利','2016股利','2017股利','2018股利','2019股利','2020股利','便宜價格','合理價格','偏高價格']]
    df.to_excel('mandy.xlsx', index=False)
# clean
for fname in range(0,7):
    if os.path.exists(str(fname)+".csv"):
        os.remove(str(fname)+".csv")

print('Finish !')